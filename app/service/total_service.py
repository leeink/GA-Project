import os
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, extract, and_
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

async def generate_excel_report(db: AsyncSession, target_year: str, save_path: str):
    from model.sales_record import SalesRecord 
    years = [str(y) for y in range(int(target_year), int(target_year)-11, -1)]
    all_data = {}

    for y in years:
        query = select(func.sum(SalesRecord.sales_price)).where(extract('year', SalesRecord.sold_at) == int(y))
        result = await db.execute(query)
        total_sales = float(result.scalar() or 0)
        
        r = {
            "제품매출": 1.0, "상품매출": 0.0, "재료비": 0.35, "노무비": 0.08, "제조경비": 0.07, 
            "상품매입원가": 0.0, "급여": 0.05, "퇴직급여": 0.005, "복리후생비": 0.01, 
            "지급임차료": 0.02, "접대비": 0.005, "감가상각비": 0.01, "무형자산상각비": 0.005, 
            "세금과공과": 0.005, "광고선전비": 0.03, "연구비": 0.005, "대손상각비": 0.005,
            "이자수익": 0.005, "배당금수익": 0.005, "임대료": 0.005, "이자비용": 0.01, "법인세비용": 0.066
        }
        st = {k: int(total_sales * v) for k, v in r.items()}
        st["매출액"] = st["제품매출"] + st["상품매출"]
        st["매출원가"] = st["재료비"] + st["노무비"] + st["제조경비"] + st["상품매입원가"]
        st["매출총이익"] = st["매출액"] - st["매출원가"]
        st["판관비"] = sum([st.get(k, 0) for k in ["급여", "퇴직급여", "복리후생비", "지급임차료", "접대비", "감가상각비", "무형자산상각비", "세금과공과", "광고선전비", "연구비", "대손상각비"]])
        st["영업이익"] = st["매출총이익"] - st["판관비"]
        st["영업외수익"] = st["이자수익"] + st["배당금수익"] + st["임대료"]
        st["법인세비용차감전이익"] = st["영업이익"] + st["영업외수익"] - st["이자비용"]
        st["당기순이익"] = st["법인세비용차감전이익"] - st["법인세비용"]
        all_data[y] = st

    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "포괄손익계산서"
    ws1.append(["대분류", "상세_계정과목"] + years)

    struct = [
        ("매출액", "제품매출", "제품매출"), ("매출액", "상품매출", "상품매출"),
        ("매출원가", "제조원가(재료비)", "재료비"), ("매출원가", "제조원가(노무비)", "노무비"),
        ("매출원가", "제조원가(제조경비)", "제조경비"), ("매출원가", "상품매입원가", "상품매입원가"),
        ("매출총이익", "매출총이익", "매출총이익"),
        ("판매비와관리비", "급여", "급여"), ("판매비와관리비", "퇴직급여", "퇴직급여"),
        ("판매비와관리비", "복리후생비", "복리후생비"), ("판매비와관리비", "지급임차료", "지급임차료"),
        ("판매비와관리비", "접대비", "접대비"), ("판매비와관리비", "감가상각비", "감가상각비"),
        ("판매비와관리비", "무형자산상각비", "무형자산상각비"), ("판매비와관리비", "세금과공과", "세금과공과"),
        ("판매비와관리비", "광고선전비", "광고선전비"), ("판매비와관리비", "연구비", "연구비"),
        ("판매비와관리비", "대손상각비", "대손상각비"),
        ("영업이익", "영업이익", "영업이익"), ("영업외손익", "이자비용", "이자비용"),
        ("당기순이익", "당기순이익", "당기순이익")
    ]
    for main, sub, key in struct:
        row = [main, sub]
        for y in years: row.append(all_data[y].get(key, 0))
        ws1.append(row)
    
    apply_style_and_width(ws1)
    wb.save(save_path)


async def generate_custom_range_report(db: AsyncSession, s_year: int, e_year: int, s_month: int, e_month: int, save_path: str):
    from model.sales_record import SalesRecord 
    
    min_y, max_y = min(s_year, e_year), max(s_year, e_year)
    min_m, max_m = min(s_month, e_month), max(s_month, e_month)
    
    ws1_years = [str(y) for y in range(max_y, min_y - 1, -1)]
    ws2_years = [str(max_y), str(max_y - 1)]
    
    fetch_years = list(set(ws1_years + ws2_years))
    fetch_years.sort(reverse=True)
    
    all_data = {}

    for y in fetch_years:
        query = select(func.sum(SalesRecord.sales_price)).where(
            and_(
                extract('year', SalesRecord.sold_at) == int(y),
                extract('month', SalesRecord.sold_at) >= min_m,
                extract('month', SalesRecord.sold_at) <= max_m
            )
        )
        result = await db.execute(query)
        total_sales = float(result.scalar() or 0)
        
        r = {
            "제품매출": 1.0, "상품매출": 0.0, "재료비": 0.35, "노무비": 0.08, "제조경비": 0.07, 
            "상품매입원가": 0.0, "급여": 0.05, "퇴직급여": 0.005, "복리후생비": 0.01, 
            "지급임차료": 0.02, "접대비": 0.005, "감가상각비": 0.01, "무형자산상각비": 0.005, 
            "세금과공과": 0.005, "광고선전비": 0.03, "연구비": 0.005, "대손상각비": 0.005,
            "이자수익": 0.005, "배당금수익": 0.005, "임대료": 0.005, "이자비용": 0.01, "법인세비용": 0.066
        }
        st = {k: int(total_sales * v) for k, v in r.items()}
        st["매출액"] = st["제품매출"] + st["상품매출"]
        st["매출원가"] = st["재료비"] + st["노무비"] + st["제조경비"] + st["상품매입원가"]
        st["매출총이익"] = st["매출액"] - st["매출원가"]
        st["판관비"] = sum([st.get(k, 0) for k in ["급여", "퇴직급여", "복리후생비", "지급임차료", "접대비", "감가상각비", "무형자산상각비", "세금과공과", "광고선전비", "연구비", "대손상각비"]])
        st["영업이익"] = st["매출총이익"] - st["판관비"]
        st["영업외수익"] = st["이자수익"] + st["배당금수익"] + st["임대료"]
        st["법인세비용차감전이익"] = st["영업이익"] + st["영업외수익"] - st["이자비용"]
        st["당기순이익"] = st["법인세비용차감전이익"] - st["법인세비용"]
        all_data[y] = st

    wb = Workbook()
    
    ws1 = wb.active
    ws1.title = "포괄손익계산서"
    m_suffix = "" if (min_m == 1 and max_m == 12) else f"({min_m}~{max_m}월)"
    ws1.append(["대분류", "상세_계정과목"] + [f"{y}년도{m_suffix}" for y in ws1_years])

    struct = [
        ("매출액", "제품매출", "제품매출"), ("매출액", "상품매출", "상품매출"),
        ("매출원가", "제조원가(재료비)", "재료비"), ("매출원가", "제조원가(노무비)", "노무비"),
        ("매출원가", "제조원가(제조경비)", "제조경비"), ("매출원가", "상품매입원가", "상품매입원가"),
        ("매출총이익", "매출총이익", "매출총이익"),
        ("판매비와관리비", "급여", "급여"), ("판매비와관리비", "퇴직급여", "퇴직급여"),
        ("판매비와관리비", "복리후생비", "복리후생비"), ("판매비와관리비", "지급임차료", "지급임차료"),
        ("판매비와관리비", "접대비", "접대비"), ("판매비와관리비", "감가상각비", "감가상각비"),
        ("판매비와관리비", "무형자산상각비", "무형자산상각비"), ("판매비와관리비", "세금과공과", "세금과공과"),
        ("판매비와관리비", "광고선전비", "광고선전비"), ("판매비와관리비", "연구비", "연구비"),
        ("판매비와관리비", "대손상각비", "대손상각비"),
        ("영업이익", "영업이익", "영업이익"), ("영업외손익", "이자비용", "이자비용"),
        ("당기순이익", "당기순이익", "당기순이익")
    ]
    for main, sub, key in struct:
        row = [main, sub]
        for y in ws1_years: row.append(all_data[y].get(key, 0))
        ws1.append(row)
    apply_style_and_width(ws1)

    ws2 = wb.create_sheet("사업보고서")
    setup_pro_business_sheet(ws2, max_y, min_m, max_m, all_data)
    apply_style_and_width(ws2)

    wb.save(save_path)


def setup_pro_business_sheet(ws2, max_y, min_m, max_m, all_data):
    gray_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    max_col = 'F' 

    ws2.merge_cells(f'A2:{max_col}2')
    ws2['A2'] = "정 기 공 시 (사업보고서)"

    ws2.merge_cells(f'A4:{max_col}4')
    title_period = f"({max_y}년도 기준)" if (min_m == 1 and max_m == 12) else f"({max_y}년도 / {min_m}~{max_m}월 기준)"
    ws2['A4'] = f"사 업 보 고 서 {title_period}"
    ws2['A4'].font = Font(size=16, bold=True)

    ws2.merge_cells('A7:A13')
    ws2['A7'] = "①\n기 업\n현 황"
    ws2['A7'].fill = gray_fill
    ws2['B7'] = "회사명"
    ws2.merge_cells('C7:D7')
    ws2['C7'] = "글로벌냉동식품"
    ws2['E7'] = "대표자"
    ws2['F7'] = "홍길동"
    ws2['B9'] = "본점소재지"
    ws2.merge_cells('C9:F9')
    ws2['C9'] = "경기도 화성시"
    ws2['B11'] = "주요사업"
    ws2.merge_cells('C11:F11')
    ws2['C11'] = "냉동식품, 신선식품 제조 및 도소매업"
    ws2['B13'] = "생산거점"
    ws2.merge_cells('C13:F13')
    ws2['C13'] = "화성시 소재 제1공장, 안산시 소재 제2공장"

    for cell in ['B7', 'E7', 'B9', 'B11', 'B13']:
        ws2[cell].fill = gray_fill

    ws2.merge_cells('A16:A22')
    ws2['A16'] = "②\n생 산\n및\n사 업"
    ws2['A16'].fill = gray_fill
    ws2['B16'] = "주력 제품군"
    ws2.merge_cells('C16:F16')
    ws2['C16'] = "삼겹살구이, 낙지볶음 도시락, 소불고기 덮밥, 치킨 도리아, 새우볶음밥, 치즈볼, 오징어 링, 감자튀김"
    ws2['B18'] = "당기 생산실적"
    ws2.merge_cells('C18:D18')
    ws2['C18'] = "1,250 톤(ton)"
    ws2['E18'] = "공장 가동률"
    ws2['F18'] = "92.5%"
    ws2['B20'] = "주요 원재료"
    ws2.merge_cells('C20:F20')
    ws2['C20'] = "국내산 돈육, 밀가루, 신선 야채류 등"
    ws2['B22'] = "연구개발(R&D)"
    ws2.merge_cells('C22:F22')
    ws2['C22'] = "당기 R&D 투자 총액 : 150,000,000 원"

    for cell in ['B16', 'B18', 'E18', 'B20', 'B22']:
        ws2[cell].fill = gray_fill

    ws2.merge_cells('A25:A31')
    ws2['A25'] = "③\n재 무\n현 황"
    ws2['A25'].fill = gray_fill
    ws2['B25'] = "구분"
    ws2['B25'].fill = gray_fill

    m_suffix = "" if (min_m == 1 and max_m == 12) else f" ({min_m}~{max_m}월)"

    ws2.merge_cells('C25:D25')
    ws2['C25'] = f"당기 ({max_y}년도){m_suffix}"
    ws2['C25'].fill = gray_fill
    ws2['D25'].fill = gray_fill

    ws2.merge_cells('E25:F25')
    ws2['E25'] = f"전기 ({max_y-1}년도){m_suffix}"
    ws2['E25'].fill = gray_fill
    ws2['F25'].fill = gray_fill

    metrics = [("매출액", 27), ("영업이익", 29), ("당기순이익", 31)]
    for key, r_idx in metrics:
        ws2[f'B{r_idx}'] = key
        ws2[f'B{r_idx}'].fill = gray_fill
        ws2.merge_cells(f'C{r_idx}:D{r_idx}')
        ws2[f'C{r_idx}'] = all_data[str(max_y)].get(key, 0)
        ws2.merge_cells(f'E{r_idx}:F{r_idx}')
        ws2[f'E{r_idx}'] = all_data[str(max_y-1)].get(key, 0)

    ws2.merge_cells('A34:A38')
    ws2['A34'] = "④\nE S G\n및\n인 력"
    ws2['A34'].fill = gray_fill
    ws2['B34'] = "총 임직원 수"
    ws2.merge_cells('C34:D34')
    ws2['C34'] = "120 명"
    ws2['E34'] = "1인평균급여"
    ws2['F34'] = "45,000,000 원"
    ws2['B36'] = "친환경 포장재"
    ws2.merge_cells('C36:D36')
    ws2['C36'] = "전 제품 적용률 : 85%"
    ws2['E36'] = "식품안전인증"
    ws2['F36'] = "☑ 스마트 HACCP 인증 완료"
    ws2['B38'] = "비고"
    ws2.merge_cells('C38:F38')

    for cell in ['B34', 'E34', 'B36', 'E36', 'B38']:
        ws2[cell].fill = gray_fill


# ==========================================
# 공통 스타일 적용 함수 (정렬, 테두리, 너비 관리)
# ==========================================
def apply_style_and_width(ws):
    side = Side(style='thin')
    border = Border(left=side, right=side, top=side, bottom=side)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right_align = Alignment(horizontal='right', vertical='center')

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if ws.title == "포괄손익계산서" and cell.row == 1:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            
            if cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.alignment = right_align
                else:
                    if ws.title == "사업보고서":
                        cell.alignment = center_align
                    else:
                        if cell.column <= 2 or cell.row == 1:
                            cell.alignment = center_align
                        else:
                            cell.alignment = right_align

            if ws.title == "사업보고서":
                if cell.row in list(range(7, 14)) + list(range(16, 23)) + list(range(25, 32)) + list(range(34, 39)):
                    if cell.column <= 6:
                        cell.border = border
            else:
                if cell.value is not None:
                    cell.border = border

    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = ws.page_margins.top = ws.page_margins.bottom = 0.5

    if ws.title == "포괄손익계산서":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        for col_idx in range(3, ws.max_column + 1):
            ws.column_dimensions[chr(64 + col_idx)].width = 16
    else: 
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.column_dimensions['A'].width = 9
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14
        # 감자튀김 안 잘리게 너비를 46으로 빵빵하게 확보!
        ws.column_dimensions['F'].width = 46